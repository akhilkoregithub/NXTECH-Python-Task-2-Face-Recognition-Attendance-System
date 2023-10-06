import glob
import face_recognition
import numpy as np
from datetime import datetime
import cv2
import os
import openpyxl
import tkinter as tk
from PIL import Image, ImageTk  

now = datetime.now()

dtString = now.strftime("%H:%M")

cap = cv2.VideoCapture(0)
FONT = cv2.FONT_HERSHEY_COMPLEX
images = []
names = []

path = r'C:\Users\kore akhil\Downloads\facereco-attendance-Excel-main\images\'*.*'
for file in glob.glob(path):
    image = cv2.imread(file)
    a = os.path.basename(file)
    b = os.path.splitext(a)[0]
    names.append(b)
    images.append(image)


def create_or_open_workbook():
    current_date = datetime.now().strftime("%Y-%m-%d")
    try:
        workbook = openpyxl.load_workbook(f"{current_date}.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.save(f"{current_date}.xlsx")

    sheet = workbook.active
    return workbook, sheet


def is_username_unique(sheet, username):
    for row in sheet.iter_rows(values_only=True):
        if row[0] == username:
            return False
    return True


def is_late_or_ontime(current_time_str, target_time_str, lateness_threshold_minutes):
    try:
        current_time = datetime.strptime(current_time_str, '%H:%M')
        target_time = datetime.strptime(target_time_str, '%H:%M')

        time_difference = current_time - target_time

        lateness_minutes = time_difference.total_seconds() / 60

        if lateness_minutes >= lateness_threshold_minutes:
            return "Late"
        else:
            return "On Time"
    except ValueError:
        return "Invalid time format. Use 'HH:MM' format (e.g., '13:45')."


def main(user_input):
    workbook, sheet = create_or_open_workbook()

    if sheet.max_row == 1:
        sheet.append(["Username", "Current Date", "Current Time", "Mark"])

    user_data = set()

    if is_username_unique(sheet, user_input):
        user_data.add(user_input)
        current_time_str = dtString
        target_time_str = '12:00'
        lateness_threshold_minutes = 6
        result = is_late_or_ontime(current_time_str, target_time_str, lateness_threshold_minutes)
        print(result)
        current_date = datetime.now().strftime("%Y-%m-%d")
        current_time = datetime.now().strftime("%H:%M:%S")
        sheet.append([user_input, current_date, current_time, result])
        workbook.save(f"{current_date}.xlsx")
    else:
        print("Username already exists. Please enter a new username.")


def encoding1(images):
    encode = []

    for img in images:
        unk_encoding = face_recognition.face_encodings(img)[0]
        encode.append(unk_encoding)
    return encode


encodelist = encoding1(images)


# Function to exit the application
def exit_app():
    cap.release()
    cv2.destroyAllWindows()
    root.quit()


# Create the main Tkinter window
root = tk.Tk()
root.title("Face Recognition")

# Create a label to display frames
label = tk.Label(root)
label.pack()

# Create a function to capture and update frames
def show_frames():
    ret, frame = cap.read()
    frame1 = cv2.resize(frame, (0, 0), None, 0.25, 0.25)
    face_locations = face_recognition.face_locations(frame1)
    curframe_encoding = face_recognition.face_encodings(frame1, face_locations)
    for encodeface, facelocation in zip(curframe_encoding, face_locations):
        results = face_recognition.compare_faces(encodelist, encodeface)
        distance = face_recognition.face_distance(encodelist, encodeface)
        match_index = np.argmin(distance)
        name = names[match_index]
        main(name)

        x1, y1, x2, y2 = facelocation
        x1, y1, x2, y2 = x1 * 4, y1 * 4, x2 * 4, y2 * 4
        cv2.rectangle(frame, (y1, x1), (y2, x2), (0, 0, 255), 3)
        cv2.putText(frame, name, (y2 + 6, x2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 0, 255), 2)

    cv2image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    img = Image.fromarray(cv2image)
    imgtk = ImageTk.PhotoImage(image=img)
    label.imgtk = imgtk
    label.configure(image=imgtk)

    label.after(10, show_frames)


# Start capturing and updating frames
show_frames()

# Create and configure the exit button
exit_button = tk.Button(root, text="Exit", command=lambda: exit_app(), bg="red", fg="white", width=10)
exit_button.pack(side="bottom", pady=10)  # Place at the bottom

# Run the Tkinter main loop
root.mainloop()
